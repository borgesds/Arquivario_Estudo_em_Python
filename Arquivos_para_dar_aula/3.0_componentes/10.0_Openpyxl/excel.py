import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('/home/borges/Documentos/Repos/arquivario/Arquivario_Estudo_em_Python/Arquivos_para_dar_aula/3.0_componentes/10.0_Openpyxl/pedidos.xlsx')

nome_palnilha = pedidos.sheetnames
planilha1 = pedidos['Página1']

"""
for linha in planilha1:

    print(linha[0].value, linha[1].value, linha[2].value)
"""

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

planilha1['B3'].value = 2200
pedidos.save('nome_palnilha.xlsx')
